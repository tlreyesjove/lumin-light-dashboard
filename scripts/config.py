"""
Central place for every assumption that feeds the synthetic data generator.

Nothing in here is "real" — it's all made up to produce a plausible-looking
B2B solar lighting business. If a number here looks wrong, change it here
and re-run the generator scripts; nothing else needs to change.
"""

import datetime
import os
import openpyxl

# ---------------------------------------------------------------------------
# Revenue targets — read directly from the financial model, not duplicated
# here. The financial model (Assumptions tab) is the single source of truth
# for what Lumin Light budgeted each year; if it changes, regenerating data
# picks up the new numbers automatically instead of silently drifting apart.
# ---------------------------------------------------------------------------
_MODEL_PATH = os.path.join(os.path.dirname(__file__), "..", "Lumin Light Financial Model - 2025-2026.xlsx")


def _read_revenue_targets():
    wb = openpyxl.load_workbook(_MODEL_PATH, data_only=True)
    ws = wb["Assumptions"]
    return {
        2025: {"Lumin Light USA": ws["C6"].value, "Lumin Light Nigeria": ws["C7"].value},
        2026: {"Lumin Light USA": ws["D6"].value, "Lumin Light Nigeria": ws["D7"].value},
    }


def read_monthly_budget():
    """Returns {(subsidiary, "YYYY-MM"): {"budget_revenue": x, "budget_ebit": y}},
    read straight from the financial model's USA PL / Nigeria PL tabs (Revenue
    row 6, EBIT row 30 — see build_financial_model.py's ROW_REVENUE/ROW_EBIT).
    Used by generate_finance.py so the Finance tab's budget columns are a real
    plan pulled from the model, not noise layered on top of actuals."""
    wb = openpyxl.load_workbook(_MODEL_PATH, data_only=True)
    sheet_by_subsidiary = {"Lumin Light USA": "USA PL", "Lumin Light Nigeria": "Nigeria PL"}
    year_blocks = [(2025, 4), (2026, 17)]  # (year, first_month_column)
    REVENUE_ROW, EBIT_ROW = 6, 30

    budget = {}
    for subsidiary, sheet_name in sheet_by_subsidiary.items():
        ws = wb[sheet_name]
        for year, first_col in year_blocks:
            for i in range(12):
                col = first_col + i
                month_str = f"{year}-{i + 1:02d}"
                budget[(subsidiary, month_str)] = {
                    "budget_revenue": ws.cell(row=REVENUE_ROW, column=col).value,
                    "budget_ebit": ws.cell(row=EBIT_ROW, column=col).value,
                }
    return budget


def read_annual_budget():
    """Returns {(subsidiary, year): {"revenue": x, "ebit": y}} — the FULL
    fiscal year's budget (the "Annual" column on each PL tab: C for 2025,
    P for 2026), not a sum of only the months that have happened so far.

    EBIT in particular should be compared against the full-year budget,
    not a prorated slice of it — a YTD actual EBIT vs. a YTD slice of
    budget EBIT double-counts the seasonality assumption (fixed costs
    aren't spread evenly the way a straight-line proration assumes), which
    is what made an early build of this look like a 150% budget overshoot
    when the real story was closer to "75% of the full year's plan banked
    with a third of the year left to go" — a much more reasonable read."""
    wb = openpyxl.load_workbook(_MODEL_PATH, data_only=True)
    sheet_by_subsidiary = {"Lumin Light USA": "USA PL", "Lumin Light Nigeria": "Nigeria PL"}
    annual_col = {2025: 3, 2026: 16}  # C, P
    REVENUE_ROW, EBIT_ROW = 6, 30

    result = {}
    for subsidiary, sheet_name in sheet_by_subsidiary.items():
        ws = wb[sheet_name]
        for year, col in annual_col.items():
            result[(subsidiary, year)] = {
                "revenue": ws.cell(row=REVENUE_ROW, column=col).value,
                "ebit": ws.cell(row=EBIT_ROW, column=col).value,
            }
    return result


def read_channel_targets():
    """Returns {(subsidiary, channel): target_dollars} for FY2026 — each
    subsidiary's annual revenue target (Assumptions D6/D7) split by
    Institutional/Commercial % (the Assumptions tab's Channel Revenue
    Target Split section, C48:D49). A change to either the overall
    target or the split % flows through automatically; nothing here is
    a second copy of either number."""
    wb = openpyxl.load_workbook(_MODEL_PATH, data_only=True)
    a = wb["Assumptions"]
    annual_target_2026 = _REVENUE_TARGETS[2026]
    pct_row = {"Lumin Light USA": 48, "Lumin Light Nigeria": 49}
    result = {}
    for subsidiary, row in pct_row.items():
        result[(subsidiary, "Institutional")] = annual_target_2026[subsidiary] * a.cell(row=row, column=3).value
        result[(subsidiary, "Commercial")] = annual_target_2026[subsidiary] * a.cell(row=row, column=4).value
    return result


def read_cost_structure():
    """Building blocks for computing ACTUAL Opex the same way the financial
    model computes BUDGET Opex — fixed costs (salaries, rent, etc.) spread
    flat across months, plus commission that scales with revenue — instead
    of actuals using an unrelated flat-%-of-revenue guess. Sharing the same
    structure means Actual-vs-Budget variance comes from real differences
    (when deals actually closed, normal month-to-month cost noise), not from
    Actual and Budget quietly using two different cost models."""
    wb = openpyxl.load_workbook(_MODEL_PATH, data_only=True)
    a = wb["Assumptions"]
    hc = wb["Headcount"]

    other_opex_2025_annual = {
        "Lumin Light USA": sum(a.cell(row=r, column=3).value for r in range(23, 29)),
        "Lumin Light Nigeria": sum(a.cell(row=r, column=4).value for r in range(23, 29)),
    }
    return {
        "benefits_loading": a["C15"].value,
        "commission_pct": a["C16"].value,
        "inflation": a["C17"].value,
        "da": {"Lumin Light USA": a["C18"].value, "Lumin Light Nigeria": a["C19"].value},
        "other_opex_2025_annual": other_opex_2025_annual,
        "headcount_base": {
            "Lumin Light USA": {2025: hc["D27"].value, 2026: hc["E27"].value},
            "Lumin Light Nigeria": {2025: hc["D28"].value, 2026: hc["E28"].value},
        },
    }


_REVENUE_TARGETS = _read_revenue_targets()

# ---------------------------------------------------------------------------
# Time window
# ---------------------------------------------------------------------------
# Lumin Light's fiscal year is the calendar year (Jan-Dec). "Current period"
# is fiscal year 2026, year-to-date — CURRENT_PERIOD_END is the fiscal
# year's real end date (Dec 31), but no generator should place actual data
# past TODAY, since those months haven't happened yet. Use TODAY, not
# CURRENT_PERIOD_END, as the upper bound whenever generating real rows.
TODAY = datetime.date(2026, 8, 25)
CURRENT_PERIOD_START = datetime.date(2026, 1, 1)
CURRENT_PERIOD_END = datetime.date(2026, 12, 31)

# One full prior fiscal year (2025) so "revenue growth vs. prior period"
# has something real to compare against instead of a made-up single number.
PRIOR_PERIOD_START = datetime.date(2025, 1, 1)
PRIOR_PERIOD_END = datetime.date(2025, 12, 31)

# ---------------------------------------------------------------------------
# Company structure
# ---------------------------------------------------------------------------
SUBSIDIARIES = ["Lumin Light USA", "Lumin Light Nigeria"]

# Revenue targets, from the financial model (see _read_revenue_targets above).
# USA skews higher — HQ, holds more of the large multilateral/government
# relationships. 2026 = current fiscal year target; 2025 = prior fiscal
# year's budget, which the "prior" period's Closed Won deals are scaled to
# match exactly (see generate_sales.py) — 2025 is a settled, complete year,
# so Sales' actual revenue and the financial model's budget for that year
# should tie out precisely, not just land in the same neighborhood.
ANNUAL_REVENUE_TARGET = _REVENUE_TARGETS[2026]
PRIOR_PERIOD_REVENUE_TARGET = _REVENUE_TARGETS[2025]
TOTAL_REVENUE_TARGET = sum(ANNUAL_REVENUE_TARGET.values())  # $15M

WAREHOUSES = {
    "Lumin Light USA": {"warehouse": "Houston, TX", "subsidiary": "Lumin Light USA"},
    "Lumin Light Nigeria": {"warehouse": "Lagos, Nigeria", "subsidiary": "Lumin Light Nigeria"},
}

BUYER_TYPES = ["Government", "Distributor", "NGO", "Multilateral"]

# Client type is a coarser rollup of buyer_type, per Tatiana: Distributors
# are resellers (Commercial), everyone else buying directly for their own
# use (Government, NGO, Multilateral) is Institutional.
CLIENT_TYPE = {
    "Government": "Institutional",
    "NGO": "Institutional",
    "Multilateral": "Institutional",
    "Distributor": "Commercial",
}

# ---------------------------------------------------------------------------
# Products — the Sol line
# ---------------------------------------------------------------------------
# unit_price / unit_cost are MY assumptions (not given in the spec) — needed
# to turn a deal's total dollar value into a plausible unit quantity, and to
# get realistic margins that widen from Sol 1 to Sol 5. Flag if these feel off.
#
# deal_value_range = total order value range from the spec (this is what's
# actually sampled per deal — quantity is *derived* from it, not the other
# way around).
PRODUCTS = {
    "Sol 1": {
        "positioning": "Pocket-sized, lowest cost, highest volume — personal preparedness",
        "deal_value_range": (5_000, 150_000),
        "unit_price": 8,
        "unit_cost": 5.20,      # ~35% gross margin
        "selection_weight": 0.35,
    },
    "Sol 2": {
        "positioning": "Basic off-grid household lighting, affordable",
        "deal_value_range": (10_000, 200_000),
        "unit_price": 25,
        "unit_cost": 15.75,     # ~37% gross margin
        "selection_weight": 0.25,
    },
    "Sol 3": {
        "positioning": "Portable, with integrated phone charging",
        "deal_value_range": (15_000, 250_000),
        "unit_price": 60,
        "unit_cost": 35.40,     # ~41% gross margin
        "selection_weight": 0.20,
    },
    "Sol 4": {
        "positioning": "Modular, built for large-scale emergency deployment",
        "deal_value_range": (50_000, 750_000),
        "unit_price": 350,
        "unit_cost": 189.00,    # ~46% gross margin
        "selection_weight": 0.12,
    },
    "Sol 5": {
        "positioning": "High-efficiency, extended runtime — shelters and facilities",
        "deal_value_range": (75_000, 1_000_000),
        "unit_price": 900,
        "unit_cost": 405.00,    # ~55% gross margin
        "selection_weight": 0.08,
    },
}

# ---------------------------------------------------------------------------
# Sales pipeline
# ---------------------------------------------------------------------------
# Stage only applies to OPEN deals — mutually exclusive with status
# (Won/Lost deals have no stage; that's what status is for). Each stage's
# win probability, per Tatiana:
STAGE_PROBABILITY = {
    "Prospecting": 0.10,
    "Qualification": 0.10,
    "Proposal": 0.50,
    "Negotiation": 0.90,
}
OPEN_STAGES = list(STAGE_PROBABILITY)

# Closed Won deals are generated per subsidiary/period by drawing deals
# until their cumulative value reaches a target (see generate_sales.py) —
# not by picking a fixed deal count, which turned out to be too noisy: a
# handful of large Sol 4/5 deals landing in one period by chance could
# swing that period's total by 20%+. Targeting revenue directly keeps
# deal-by-deal randomness (product, buyer, exact value) while controlling
# the metric that actually matters (period revenue, and the growth
# between periods).
OPEN_DEALS = 90

# Base win rate for a closed deal; nudged up/down per buyer type below.
BASE_WIN_RATE = 0.32
WIN_RATE_ADJUSTMENT = {
    "Government": -0.06,   # competitive tenders, harder to win
    "Distributor": 0.10,   # repeat channel relationships
    "NGO": 0.00,
    "Multilateral": -0.02,
}

# Typical days from deal creation to expected close, by buyer type
# (government tenders move slower than distributor reorders).
SALES_CYCLE_DAYS = {
    "Government": (120, 270),
    "Distributor": (30, 90),
    "NGO": (45, 120),
    "Multilateral": (90, 200),
}

# ---------------------------------------------------------------------------
# Finance
# ---------------------------------------------------------------------------
BLENDED_OPEX_PCT_OF_REVENUE = 0.32   # opex as % of revenue, before EBIT — actuals only; budget comes from the financial model
STARTING_CASH_BALANCE = 2_200_000    # cash balance at PRIOR_PERIOD_START
AR_LAG_DAYS = 35                     # institutional buyers pay slowly — cash lags revenue

# ---------------------------------------------------------------------------
# Inventory
# ---------------------------------------------------------------------------
# Lead time = how long it takes to receive a replenishment order (e.g. an
# order of ~100,000 units) after placing it. Tatiana's assumption: 1 month,
# same for average and worst case. Kept as two separate constants (not one)
# so this still works if avg/max lead time ever need to diverge — e.g. the
# Nigeria warehouse taking longer on a bad shipping month than usual.
AVG_LEAD_TIME_MONTHS = 1
MAX_LEAD_TIME_MONTHS = 1

RANDOM_SEED = 42
