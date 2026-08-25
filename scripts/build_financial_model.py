"""
Builds "Lumin Light Financial Model - 2025-2026.xlsx" — a from-scratch
financial model (structure inspired by a real Be Girl financial model
template, but every number here is invented for Lumin Light; nothing
real carries over).

This is a one-off build script, run by hand when assumptions change —
not part of the regular synthetic-data pipeline (generate_data.py).
Every number lives in the Assumptions or Headcount tab; every other cell
is a formula, so changing an assumption and reopening the file in Excel
recalculates the whole model.

Usage:
    python3 build_financial_model.py
    python3 ../[xlsx skill]/scripts/recalc.py "../Lumin Light Financial Model - 2025-2026.xlsx"
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

OUT_PATH = os.path.join(os.path.dirname(__file__), "..", "Lumin Light Financial Model - 2025-2026.xlsx")

BLUE = Font(color="0000FF")
BLUE_BOLD = Font(color="0000FF", bold=True)
BLACK_BOLD = Font(bold=True)
GREEN = Font(color="008000")
GREEN_BOLD = Font(color="008000", bold=True)
TITLE_FONT = Font(bold=True, size=14)
SECTION_FONT = Font(bold=True, size=11)
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

USD_FMT = '$#,##0;($#,##0);"-"'
PCT_FMT = '0.0%'

MONTH_NAMES = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]


def set(ws, row, col, value, font=None, num_fmt=None, fill=None, align=None):
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if num_fmt:
        cell.number_format = num_fmt
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    return cell


def col_letter(n):
    return get_column_letter(n)


# ---------------------------------------------------------------------------
# Assumptions
# ---------------------------------------------------------------------------

def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    for c in range(3, 15):
        ws.column_dimensions[col_letter(c)].width = 13

    set(ws, 1, 2, "Lumin Light — 2025-2026 Financial Model — Assumptions", TITLE_FONT)
    set(ws, 2, 2, "Every hardcoded number in this workbook lives on this tab or the Headcount tab (blue text). Everything else is a formula.", Font(italic=True, size=9))

    set(ws, 4, 2, "REVENUE TARGETS ($)", SECTION_FONT)
    set(ws, 5, 2, "Subsidiary", BLACK_BOLD); set(ws, 5, 3, "2025", BLACK_BOLD); set(ws, 5, 4, "2026", BLACK_BOLD)
    set(ws, 6, 2, "Lumin Light USA")
    set(ws, 6, 3, 7850000, BLUE, USD_FMT, YELLOW_FILL)
    set(ws, 6, 4, 9000000, BLUE, USD_FMT, YELLOW_FILL)
    set(ws, 7, 2, "Lumin Light Nigeria")
    set(ws, 7, 3, 5100000, BLUE, USD_FMT, YELLOW_FILL)
    set(ws, 7, 4, 6000000, BLUE, USD_FMT, YELLOW_FILL)
    set(ws, 8, 2, "Total", BLACK_BOLD)
    set(ws, 8, 3, "=SUM(C6:C7)", BLACK_BOLD, USD_FMT)
    set(ws, 8, 4, "=SUM(D6:D7)", BLACK_BOLD, USD_FMT)

    set(ws, 10, 2, "COST & MARGIN ASSUMPTIONS", SECTION_FONT)
    set(ws, 11, 2, "Blended Gross Margin % (dollar-weighted across Sol 1-5 mix)")
    set(ws, 11, 3, 0.473, BLUE, PCT_FMT, YELLOW_FILL)
    set(ws, 12, 2, "Empirically calibrated from an actual generate_sales_data() run (config.RANDOM_SEED=42), not a theoretical average — see Learning Doc 2.4. If PRODUCTS or RANDOM_SEED in config.py ever change, recheck this against a fresh run's realized margin.", Font(italic=True, size=9))

    set(ws, 14, 2, "OPERATING EXPENSE ASSUMPTIONS", SECTION_FONT)
    set(ws, 15, 2, "Benefits & Payroll Tax Loading (% of base salary)")
    set(ws, 15, 3, 0.25, BLUE, PCT_FMT, YELLOW_FILL)
    set(ws, 16, 2, "Sales Commission (% of Revenue)")
    set(ws, 16, 3, 0.02, BLUE, PCT_FMT, YELLOW_FILL)
    set(ws, 17, 2, "Annual Cost Inflation (2025 -> 2026)")
    set(ws, 17, 3, 0.035, BLUE, PCT_FMT, YELLOW_FILL)
    set(ws, 18, 2, "Annual Depreciation & Amortization — USA ($)")
    set(ws, 18, 3, 40000, BLUE, USD_FMT, YELLOW_FILL)
    set(ws, 19, 2, "Annual Depreciation & Amortization — Nigeria ($)")
    set(ws, 19, 3, 20000, BLUE, USD_FMT, YELLOW_FILL)

    set(ws, 21, 2, "OTHER OPERATING EXPENSE CATEGORIES — 2025 Annual Baseline ($)", SECTION_FONT)
    set(ws, 22, 2, "Category", BLACK_BOLD); set(ws, 22, 3, "USA", BLACK_BOLD)
    set(ws, 22, 4, "Nigeria", BLACK_BOLD); set(ws, 22, 5, "Total", BLACK_BOLD)
    categories = [
        ("Rent & Facilities", 150000, 90000),
        ("Travel", 230000, 140000),
        ("Professional Services", 140000, 60000),
        ("Marketing", 90000, 40000),
        ("Freight & Logistics", 180000, 160000),
        ("G&A / Admin", 140000, 80000),
    ]
    first_cat_row = 23
    for i, (name, usa, nga) in enumerate(categories):
        r = first_cat_row + i
        set(ws, r, 2, name)
        set(ws, r, 3, usa, BLUE, USD_FMT, YELLOW_FILL)
        set(ws, r, 4, nga, BLUE, USD_FMT, YELLOW_FILL)
        set(ws, r, 5, f"=SUM(C{r}:D{r})", None, USD_FMT)
    last_cat_row = first_cat_row + len(categories) - 1
    total_row = last_cat_row + 1
    set(ws, total_row, 2, "Total", BLACK_BOLD)
    set(ws, total_row, 3, f"=SUM(C{first_cat_row}:C{last_cat_row})", BLACK_BOLD, USD_FMT)
    set(ws, total_row, 4, f"=SUM(D{first_cat_row}:D{last_cat_row})", BLACK_BOLD, USD_FMT)
    set(ws, total_row, 5, f"=SUM(E{first_cat_row}:E{last_cat_row})", BLACK_BOLD, USD_FMT)

    seas_header_row = total_row + 2
    seas_row = seas_header_row + 1
    check_row = seas_row + 1
    set(ws, seas_header_row, 2, "MONTHLY REVENUE SEASONALITY INDEX (institutional year-end budget flush pushes Q4 higher)", SECTION_FONT)
    weights = [0.85, 0.85, 0.90, 0.90, 0.95, 1.00, 0.95, 0.90, 1.00, 1.05, 1.30, 1.35]
    for i, m in enumerate(MONTH_NAMES):
        set(ws, seas_row - 1, 3 + i, m, BLACK_BOLD)
        set(ws, seas_row, 3 + i, weights[i], BLUE, "0.00", YELLOW_FILL)
    set(ws, check_row, 2, "Check (sums to 12.00 if evenly weighted; doesn't have to — formulas normalize either way)")
    set(ws, check_row, 3, f"=SUM(C{seas_row}:N{seas_row})", None, "0.00")

    ar_row = check_row + 2
    set(ws, ar_row, 2, "AR Collection Lag (months) — institutional buyers pay slowly; matches the Python actuals model")
    set(ws, ar_row, 3, 1, BLUE, "0", YELLOW_FILL)

    cash_header_row = ar_row + 2
    cash_row1 = cash_header_row + 1
    set(ws, cash_header_row, 2, "STARTING CASH BALANCE — Jan 1, 2025 ($)", SECTION_FONT)
    set(ws, cash_row1, 2, "Lumin Light USA")
    set(ws, cash_row1, 3, 2200000, BLUE, USD_FMT, YELLOW_FILL)
    set(ws, cash_row1 + 1, 2, "Lumin Light Nigeria")
    set(ws, cash_row1 + 1, 3, 2200000, BLUE, USD_FMT, YELLOW_FILL)

    return {
        "revenue_row": {"Lumin Light USA": 6, "Lumin Light Nigeria": 7},
        "gross_margin_cell": "C11",
        "benefits_loading_cell": "C15",
        "commission_cell": "C16",
        "inflation_cell": "C17",
        "da_row": {"Lumin Light USA": 18, "Lumin Light Nigeria": 19},
        "opex_cat_rows": {name: first_cat_row + i for i, (name, _, _) in enumerate(categories)},
        "opex_cat_col": {"Lumin Light USA": 3, "Lumin Light Nigeria": 4},
        "seasonality_row": seas_row,
        "ar_lag_cell": f"C{ar_row}",
        "cash_row": {"Lumin Light USA": cash_row1, "Lumin Light Nigeria": cash_row1 + 1},
    }


# ---------------------------------------------------------------------------
# Headcount
# ---------------------------------------------------------------------------

ROSTER = [
    ("CEO", "Lumin Light USA", 240000),
    ("COO", "Lumin Light USA", 190000),
    ("CFO (Fractional)", "Lumin Light USA", 100000),
    ("VP Sales", "Lumin Light USA", 180000),
    ("Regional Sales Director, West Africa", "Lumin Light Nigeria", 115000),
    ("Sales Manager, Government & Multilateral Accounts", "Lumin Light USA", 135000),
    ("Sales Manager, NGO Accounts", "Lumin Light USA", 120000),
    ("Business Development Manager", "Lumin Light Nigeria", 75000),
    ("Tender & Bid Manager", "Lumin Light USA", 100000),
    ("Supply Chain Director", "Lumin Light USA", 155000),
    ("Warehouse Manager, Houston", "Lumin Light USA", 80000),
    ("Warehouse Manager, Lagos", "Lumin Light Nigeria", 55000),
    ("Logistics & Procurement Coordinator", "Lumin Light USA", 75000),
    ("Inventory & Ops Analyst", "Lumin Light Nigeria", 48000),
    ("Controller", "Lumin Light USA", 125000),
    ("Finance & Accounting Associate", "Lumin Light Nigeria", 48000),
    ("Office & Admin Manager", "Lumin Light Nigeria", 38000),
    ("Marketing Manager", "Lumin Light USA", 95000),
    ("Customer Success / Account Manager", "Lumin Light Nigeria", 58000),
    ("IT & Systems Administrator", "Lumin Light USA", 90000),
]


def build_headcount(wb, assumptions_refs):
    ws = wb.create_sheet("Headcount")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 46
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 16
    ws.column_dimensions["E"].width = 16

    set(ws, 1, 2, "Lumin Light — Headcount & Salaries", TITLE_FONT)
    set(ws, 2, 2, "20 roles, ~20-person operation per Tatiana. Base salaries only — Assumptions!benefits loading is applied on the P&L tabs, not here.", Font(italic=True, size=9))

    set(ws, 4, 2, "Role", BLACK_BOLD); set(ws, 4, 3, "Subsidiary", BLACK_BOLD)
    set(ws, 4, 4, "2025 Annual Base Salary ($)", BLACK_BOLD)
    set(ws, 4, 5, "2026 Annual Base Salary ($)", BLACK_BOLD)

    first_row = 5
    infl_cell = f"Assumptions!${assumptions_refs['inflation_cell'][0]}${assumptions_refs['inflation_cell'][1:]}"
    for i, (role, sub, salary) in enumerate(ROSTER):
        r = first_row + i
        set(ws, r, 2, role)
        set(ws, r, 3, sub)
        set(ws, r, 4, salary, BLUE, USD_FMT, YELLOW_FILL)
        set(ws, r, 5, f"=D{r}*(1+{infl_cell})", None, USD_FMT)
    last_row = first_row + len(ROSTER) - 1

    total_row = last_row + 2
    usa_row = total_row + 1
    nga_row = total_row + 2
    set(ws, total_row, 2, "Total Base Salary", BLACK_BOLD)
    set(ws, total_row, 4, f"=SUM(D{first_row}:D{last_row})", BLACK_BOLD, USD_FMT)
    set(ws, total_row, 5, f"=SUM(E{first_row}:E{last_row})", BLACK_BOLD, USD_FMT)
    set(ws, usa_row, 2, "Lumin Light USA Total")
    set(ws, usa_row, 4, f'=SUMIF(C{first_row}:C{last_row},"Lumin Light USA",D{first_row}:D{last_row})', None, USD_FMT)
    set(ws, usa_row, 5, f'=SUMIF(C{first_row}:C{last_row},"Lumin Light USA",E{first_row}:E{last_row})', None, USD_FMT)
    set(ws, nga_row, 2, "Lumin Light Nigeria Total")
    set(ws, nga_row, 4, f'=SUMIF(C{first_row}:C{last_row},"Lumin Light Nigeria",D{first_row}:D{last_row})', None, USD_FMT)
    set(ws, nga_row, 5, f'=SUMIF(C{first_row}:C{last_row},"Lumin Light Nigeria",E{first_row}:E{last_row})', None, USD_FMT)

    return {
        "total_row": {"Lumin Light USA": usa_row, "Lumin Light Nigeria": nga_row},
    }


# ---------------------------------------------------------------------------
# Subsidiary P&L (USA PL, Nigeria PL) — same layout, different source refs
# ---------------------------------------------------------------------------

ROW_REVENUE = 6
ROW_COGS = 9
ROW_GP = 11
ROW_GM_PCT = 12
ROW_SALARIES = 15
ROW_COMMISSION = 16
ROW_RENT = 17
ROW_TRAVEL = 18
ROW_PROFSVCS = 19
ROW_MARKETING = 20
ROW_FREIGHT = 21
ROW_GA = 22
ROW_TOTAL_OPEX = 23
ROW_EBITDA = 25
ROW_EBITDA_PCT = 26
ROW_DA = 28
ROW_EBIT = 30
ROW_EBIT_PCT = 31

OPEX_CATEGORY_ROWS = {
    ROW_RENT: "Rent & Facilities",
    ROW_TRAVEL: "Travel",
    ROW_PROFSVCS: "Professional Services",
    ROW_MARKETING: "Marketing",
    ROW_FREIGHT: "Freight & Logistics",
    ROW_GA: "G&A / Admin",
}

YEAR_BLOCKS = [
    (2025, 3, 4),   # (year, annual_col, first_month_col)
    (2026, 16, 17),
]


def month_cols(first_month_col):
    return [first_month_col + i for i in range(12)]


def build_subsidiary_pl(wb, sheet_name, subsidiary, assumptions_refs, headcount_refs):
    ws = wb.create_sheet(sheet_name)
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    for c in range(3, 29):
        ws.column_dimensions[col_letter(c)].width = 11

    set(ws, 1, 2, f"{subsidiary} — P&L", TITLE_FONT)

    for year, annual_col, first_month_col in YEAR_BLOCKS:
        set(ws, 2, annual_col, f"FY{year}", SECTION_FONT)
        set(ws, 3, annual_col, "Annual", BLACK_BOLD)
        for i, m in enumerate(MONTH_NAMES):
            set(ws, 3, first_month_col + i, m, BLACK_BOLD)

    set(ws, ROW_REVENUE, 2, "Revenue", BLACK_BOLD)
    set(ws, ROW_COGS, 2, "COGS")
    set(ws, ROW_GP, 2, "Gross Profit", BLACK_BOLD)
    set(ws, ROW_GM_PCT, 2, "Gross Margin %", Font(italic=True))
    set(ws, 14, 2, "Operating Expenses", SECTION_FONT)
    set(ws, ROW_SALARIES, 2, "Salaries & Benefits")
    set(ws, ROW_COMMISSION, 2, "Sales Commission")
    for row, label in OPEX_CATEGORY_ROWS.items():
        set(ws, row, 2, label)
    set(ws, ROW_TOTAL_OPEX, 2, "Total Operating Expenses", BLACK_BOLD)
    set(ws, ROW_EBITDA, 2, "EBITDA", BLACK_BOLD)
    set(ws, ROW_EBITDA_PCT, 2, "EBITDA Margin %", Font(italic=True))
    set(ws, ROW_DA, 2, "Depreciation & Amortization")
    set(ws, ROW_EBIT, 2, "EBIT", BLACK_BOLD)
    set(ws, ROW_EBIT_PCT, 2, "EBIT Margin %", Font(italic=True))

    rev_row_assump = assumptions_refs["revenue_row"][subsidiary]
    seas_row = assumptions_refs["seasonality_row"]
    gm_cell = assumptions_refs["gross_margin_cell"]
    benefits_cell = assumptions_refs["benefits_loading_cell"]
    commission_cell = assumptions_refs["commission_cell"]
    inflation_cell = assumptions_refs["inflation_cell"]
    da_row_assump = assumptions_refs["da_row"][subsidiary]
    opex_col = assumptions_refs["opex_cat_col"][subsidiary]
    hc_total_row = headcount_refs["total_row"][subsidiary]

    for year, annual_col, first_month_col in YEAR_BLOCKS:
        rev_col_letter = "C" if year == 2025 else "D"     # Assumptions!C=2025, D=2026
        hc_col_letter = "D" if year == 2025 else "E"       # Headcount!D=2025, E=2026
        opex_inflate = "" if year == 2025 else f"*(1+Assumptions!${inflation_cell[0]}${inflation_cell[1:]})"

        for i, mc in enumerate(month_cols(first_month_col)):
            L = col_letter(mc)
            seas_col = col_letter(3 + i)  # Assumptions seasonality columns C..N

            # Revenue
            set(ws, ROW_REVENUE, mc,
                f"=Assumptions!${rev_col_letter}${rev_row_assump}*Assumptions!${seas_col}${seas_row}"
                f"/SUM(Assumptions!$C${seas_row}:$N${seas_row})", GREEN, USD_FMT)
            # COGS
            set(ws, ROW_COGS, mc, f"={L}{ROW_REVENUE}*(1-Assumptions!${gm_cell[0]}${gm_cell[1:]})", GREEN, USD_FMT)
            # Gross Profit
            set(ws, ROW_GP, mc, f"={L}{ROW_REVENUE}-{L}{ROW_COGS}", None, USD_FMT)
            set(ws, ROW_GM_PCT, mc, f"=IF({L}{ROW_REVENUE}=0,0,{L}{ROW_GP}/{L}{ROW_REVENUE})", Font(italic=True), PCT_FMT)

            # Salaries & Benefits
            set(ws, ROW_SALARIES, mc,
                f"=Headcount!${hc_col_letter}${hc_total_row}*(1+Assumptions!${benefits_cell[0]}${benefits_cell[1:]})/12",
                GREEN, USD_FMT)
            # Sales Commission
            set(ws, ROW_COMMISSION, mc, f"={L}{ROW_REVENUE}*Assumptions!${commission_cell[0]}${commission_cell[1:]}", GREEN, USD_FMT)
            # Other opex categories
            for row in OPEX_CATEGORY_ROWS:
                cat_row_assump = assumptions_refs["opex_cat_rows"][OPEX_CATEGORY_ROWS[row]]
                set(ws, row, mc,
                    f"=Assumptions!${col_letter(opex_col)}${cat_row_assump}{opex_inflate}/12",
                    GREEN, USD_FMT)
            # Total Opex
            set(ws, ROW_TOTAL_OPEX, mc, f"=SUM({L}{ROW_SALARIES}:{L}{ROW_GA})", BLACK_BOLD, USD_FMT)
            # EBITDA
            set(ws, ROW_EBITDA, mc, f"={L}{ROW_GP}-{L}{ROW_TOTAL_OPEX}", BLACK_BOLD, USD_FMT)
            set(ws, ROW_EBITDA_PCT, mc, f"=IF({L}{ROW_REVENUE}=0,0,{L}{ROW_EBITDA}/{L}{ROW_REVENUE})", Font(italic=True), PCT_FMT)
            # D&A
            set(ws, ROW_DA, mc, f"=Assumptions!$C${da_row_assump}/12", GREEN, USD_FMT)
            # EBIT
            set(ws, ROW_EBIT, mc, f"={L}{ROW_EBITDA}-{L}{ROW_DA}", BLACK_BOLD, USD_FMT)
            set(ws, ROW_EBIT_PCT, mc, f"=IF({L}{ROW_REVENUE}=0,0,{L}{ROW_EBIT}/{L}{ROW_REVENUE})", Font(italic=True), PCT_FMT)

        # Annual columns = SUM of the 12 monthly columns for $ rows; ratio for % rows
        AC = col_letter(annual_col)
        first_m = col_letter(first_month_col)
        last_m = col_letter(first_month_col + 11)
        for row in [ROW_REVENUE, ROW_COGS, ROW_GP, ROW_SALARIES, ROW_COMMISSION,
                    ROW_RENT, ROW_TRAVEL, ROW_PROFSVCS, ROW_MARKETING, ROW_FREIGHT, ROW_GA,
                    ROW_TOTAL_OPEX, ROW_EBITDA, ROW_DA, ROW_EBIT]:
            bold = row in (ROW_REVENUE, ROW_GP, ROW_TOTAL_OPEX, ROW_EBITDA, ROW_EBIT)
            set(ws, row, annual_col, f"=SUM({first_m}{row}:{last_m}{row})", BLACK_BOLD if bold else None, USD_FMT)
        for row in [ROW_GM_PCT, ROW_EBITDA_PCT, ROW_EBIT_PCT]:
            num_row = {ROW_GM_PCT: ROW_GP, ROW_EBITDA_PCT: ROW_EBITDA, ROW_EBIT_PCT: ROW_EBIT}[row]
            set(ws, row, annual_col, f"=IF({AC}{ROW_REVENUE}=0,0,{AC}{num_row}/{AC}{ROW_REVENUE})", Font(italic=True), PCT_FMT)

    return {
        "rows": dict(revenue=ROW_REVENUE, cogs=ROW_COGS, gp=ROW_GP, total_opex=ROW_TOTAL_OPEX,
                     ebitda=ROW_EBITDA, da=ROW_DA, ebit=ROW_EBIT,
                     gm_pct=ROW_GM_PCT, ebitda_pct=ROW_EBITDA_PCT, ebit_pct=ROW_EBIT_PCT,
                     salaries=ROW_SALARIES, commission=ROW_COMMISSION, **{v: k for k, v in OPEX_CATEGORY_ROWS.items()}),
    }


# ---------------------------------------------------------------------------
# Consolidated P&L — sum of USA PL + Nigeria PL (no intercompany elimination
# needed: both subsidiaries buy at the same price from an external supplier,
# unlike the real Be Girl model where MOZ buys from INC at a markup)
# ---------------------------------------------------------------------------

def build_consolidated_pl(wb):
    ws = wb.create_sheet("Consolidated PL")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 30
    for c in range(3, 29):
        ws.column_dimensions[col_letter(c)].width = 11

    set(ws, 1, 2, "Lumin Light — Consolidated P&L (USA + Nigeria)", TITLE_FONT)
    set(ws, 2, 2, "No intercompany elimination line — both subsidiaries buy at the same price from an external supplier, so Consolidated is a straight sum.", Font(italic=True, size=9))

    for year, annual_col, first_month_col in YEAR_BLOCKS:
        set(ws, 3, annual_col, "Annual", BLACK_BOLD)
        for i, m in enumerate(MONTH_NAMES):
            set(ws, 3, first_month_col + i, m, BLACK_BOLD)
        set(ws, 2, annual_col, f"FY{year}", SECTION_FONT)

    set(ws, ROW_REVENUE, 2, "Revenue", BLACK_BOLD)
    set(ws, ROW_COGS, 2, "COGS")
    set(ws, ROW_GP, 2, "Gross Profit", BLACK_BOLD)
    set(ws, ROW_GM_PCT, 2, "Gross Margin %", Font(italic=True))
    set(ws, 14, 2, "Operating Expenses", SECTION_FONT)
    set(ws, ROW_SALARIES, 2, "Salaries & Benefits")
    set(ws, ROW_COMMISSION, 2, "Sales Commission")
    for row, label in OPEX_CATEGORY_ROWS.items():
        set(ws, row, 2, label)
    set(ws, ROW_TOTAL_OPEX, 2, "Total Operating Expenses", BLACK_BOLD)
    set(ws, ROW_EBITDA, 2, "EBITDA", BLACK_BOLD)
    set(ws, ROW_EBITDA_PCT, 2, "EBITDA Margin %", Font(italic=True))
    set(ws, ROW_DA, 2, "Depreciation & Amortization")
    set(ws, ROW_EBIT, 2, "EBIT", BLACK_BOLD)
    set(ws, ROW_EBIT_PCT, 2, "EBIT Margin %", Font(italic=True))

    dollar_rows = [ROW_REVENUE, ROW_COGS, ROW_GP, ROW_SALARIES, ROW_COMMISSION,
                   ROW_RENT, ROW_TRAVEL, ROW_PROFSVCS, ROW_MARKETING, ROW_FREIGHT, ROW_GA,
                   ROW_TOTAL_OPEX, ROW_EBITDA, ROW_DA, ROW_EBIT]
    bold_rows = {ROW_REVENUE, ROW_GP, ROW_TOTAL_OPEX, ROW_EBITDA, ROW_EBIT}
    pct_rows = {ROW_GM_PCT: ROW_GP, ROW_EBITDA_PCT: ROW_EBITDA, ROW_EBIT_PCT: ROW_EBIT}

    for year, annual_col, first_month_col in YEAR_BLOCKS:
        cols = [annual_col] + month_cols(first_month_col)
        for c in cols:
            L = col_letter(c)
            for row in dollar_rows:
                set(ws, row, c, f"='USA PL'!{L}{row}+'Nigeria PL'!{L}{row}",
                    BLACK_BOLD if row in bold_rows else None, USD_FMT)
            for pct_row, num_row in pct_rows.items():
                set(ws, pct_row, c, f"=IF({L}{ROW_REVENUE}=0,0,{L}{num_row}/{L}{ROW_REVENUE})", Font(italic=True), PCT_FMT)


# ---------------------------------------------------------------------------
# Cash Flow — continuous 24-month rollforward (Jan 2025 - Dec 2026),
# USA / Nigeria / Consolidated blocks
# ---------------------------------------------------------------------------

CF_ROW_OPEN = {"Lumin Light USA": 6, "Lumin Light Nigeria": 14}
CF_ROW_COLLECTED = {"Lumin Light USA": 7, "Lumin Light Nigeria": 15}
CF_ROW_COGS = {"Lumin Light USA": 8, "Lumin Light Nigeria": 16}
CF_ROW_OPEX = {"Lumin Light USA": 9, "Lumin Light Nigeria": 17}
CF_ROW_NET = {"Lumin Light USA": 10, "Lumin Light Nigeria": 18}
CF_ROW_CLOSE = {"Lumin Light USA": 11, "Lumin Light Nigeria": 19}
CONSOL_OPEN, CONSOL_COLLECTED, CONSOL_COGS, CONSOL_OPEX, CONSOL_NET, CONSOL_CLOSE = 22, 23, 24, 25, 26, 27


def pl_col_for_month_index(i):
    """Absolute month index 0-23 (Jan25..Dec26) -> the matching PL tab column letter."""
    return col_letter(4 + i) if i < 12 else col_letter(17 + (i - 12))


def build_cash_flow(wb, assumptions_refs):
    ws = wb.create_sheet("Cash Flow")
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 22
    for c in range(3, 27):
        ws.column_dimensions[col_letter(c)].width = 11

    set(ws, 1, 2, "Lumin Light — Cash Flow", TITLE_FONT)
    set(ws, 2, 2, "1-month AR collection lag on Revenue (institutional buyers pay slowly); COGS and Opex paid same month — same logic as the Python actuals model.", Font(italic=True, size=9))

    month_labels = [f"{m}-25" for m in MONTH_NAMES] + [f"{m}-26" for m in MONTH_NAMES]
    for i, label in enumerate(month_labels):
        set(ws, 4, 3 + i, label, BLACK_BOLD)

    set(ws, 5, 2, "Lumin Light USA", SECTION_FONT)
    set(ws, 13, 2, "Lumin Light Nigeria", SECTION_FONT)
    set(ws, 21, 2, "Consolidated", SECTION_FONT)
    row_labels = {"open": "Opening Balance", "collected": "Revenue Collected",
                  "cogs": "COGS Paid", "opex": "Opex Paid", "net": "Net Cash Change", "close": "Closing Balance"}
    for sub, rows in [("Lumin Light USA", (CF_ROW_OPEN, CF_ROW_COLLECTED, CF_ROW_COGS, CF_ROW_OPEX, CF_ROW_NET, CF_ROW_CLOSE)),
                       ("Lumin Light Nigeria", (CF_ROW_OPEN, CF_ROW_COLLECTED, CF_ROW_COGS, CF_ROW_OPEX, CF_ROW_NET, CF_ROW_CLOSE))]:
        pass
    set(ws, CF_ROW_OPEN["Lumin Light USA"], 2, row_labels["open"])
    set(ws, CF_ROW_COLLECTED["Lumin Light USA"], 2, row_labels["collected"])
    set(ws, CF_ROW_COGS["Lumin Light USA"], 2, row_labels["cogs"])
    set(ws, CF_ROW_OPEX["Lumin Light USA"], 2, row_labels["opex"])
    set(ws, CF_ROW_NET["Lumin Light USA"], 2, row_labels["net"], BLACK_BOLD)
    set(ws, CF_ROW_CLOSE["Lumin Light USA"], 2, row_labels["close"], BLACK_BOLD)
    set(ws, CF_ROW_OPEN["Lumin Light Nigeria"], 2, row_labels["open"])
    set(ws, CF_ROW_COLLECTED["Lumin Light Nigeria"], 2, row_labels["collected"])
    set(ws, CF_ROW_COGS["Lumin Light Nigeria"], 2, row_labels["cogs"])
    set(ws, CF_ROW_OPEX["Lumin Light Nigeria"], 2, row_labels["opex"])
    set(ws, CF_ROW_NET["Lumin Light Nigeria"], 2, row_labels["net"], BLACK_BOLD)
    set(ws, CF_ROW_CLOSE["Lumin Light Nigeria"], 2, row_labels["close"], BLACK_BOLD)
    set(ws, CONSOL_OPEN, 2, row_labels["open"])
    set(ws, CONSOL_COLLECTED, 2, row_labels["collected"])
    set(ws, CONSOL_COGS, 2, row_labels["cogs"])
    set(ws, CONSOL_OPEX, 2, row_labels["opex"])
    set(ws, CONSOL_NET, 2, row_labels["net"], BLACK_BOLD)
    set(ws, CONSOL_CLOSE, 2, row_labels["close"], BLACK_BOLD)

    cash_row = assumptions_refs["cash_row"]
    pl_sheet = {"Lumin Light USA": "'USA PL'", "Lumin Light Nigeria": "'Nigeria PL'"}

    for sub in ["Lumin Light USA", "Lumin Light Nigeria"]:
        open_r, coll_r, cogs_r, opex_r, net_r, close_r = (
            CF_ROW_OPEN[sub], CF_ROW_COLLECTED[sub], CF_ROW_COGS[sub], CF_ROW_OPEX[sub], CF_ROW_NET[sub], CF_ROW_CLOSE[sub])
        start_cash_row = cash_row[sub]
        for i in range(24):
            c = 3 + i
            L = col_letter(c)
            pl_col = pl_col_for_month_index(i)

            # Opening balance: first month pulls from Assumptions; every other month = prior month's closing
            if i == 0:
                set(ws, open_r, c, f"=Assumptions!$C${start_cash_row}", GREEN, USD_FMT)
            else:
                prev_L = col_letter(c - 1)
                set(ws, open_r, c, f"={prev_L}{close_r}", None, USD_FMT)

            # Revenue collected: 1-month lag (first month has no prior, uses same month)
            if i == 0:
                set(ws, coll_r, c, f"={pl_sheet[sub]}!{pl_col}{ROW_REVENUE}", GREEN, USD_FMT)
            else:
                prev_pl_col = pl_col_for_month_index(i - 1)
                set(ws, coll_r, c, f"={pl_sheet[sub]}!{prev_pl_col}{ROW_REVENUE}", GREEN, USD_FMT)

            set(ws, cogs_r, c, f"={pl_sheet[sub]}!{pl_col}{ROW_COGS}", GREEN, USD_FMT)
            set(ws, opex_r, c, f"={pl_sheet[sub]}!{pl_col}{ROW_TOTAL_OPEX}", GREEN, USD_FMT)
            set(ws, net_r, c, f"={L}{coll_r}-{L}{cogs_r}-{L}{opex_r}", BLACK_BOLD, USD_FMT)
            set(ws, close_r, c, f"={L}{open_r}+{L}{net_r}", BLACK_BOLD, USD_FMT)

    for i in range(24):
        c = 3 + i
        L = col_letter(c)
        set(ws, CONSOL_OPEN, c, f"={L}{CF_ROW_OPEN['Lumin Light USA']}+{L}{CF_ROW_OPEN['Lumin Light Nigeria']}", None, USD_FMT)
        set(ws, CONSOL_COLLECTED, c, f"={L}{CF_ROW_COLLECTED['Lumin Light USA']}+{L}{CF_ROW_COLLECTED['Lumin Light Nigeria']}", None, USD_FMT)
        set(ws, CONSOL_COGS, c, f"={L}{CF_ROW_COGS['Lumin Light USA']}+{L}{CF_ROW_COGS['Lumin Light Nigeria']}", None, USD_FMT)
        set(ws, CONSOL_OPEX, c, f"={L}{CF_ROW_OPEX['Lumin Light USA']}+{L}{CF_ROW_OPEX['Lumin Light Nigeria']}", None, USD_FMT)
        set(ws, CONSOL_NET, c, f"={L}{CF_ROW_NET['Lumin Light USA']}+{L}{CF_ROW_NET['Lumin Light Nigeria']}", BLACK_BOLD, USD_FMT)
        set(ws, CONSOL_CLOSE, c, f"={L}{CF_ROW_CLOSE['Lumin Light USA']}+{L}{CF_ROW_CLOSE['Lumin Light Nigeria']}", BLACK_BOLD, USD_FMT)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    wb = Workbook()
    wb.remove(wb.active)  # remove default blank sheet

    assumptions_refs = build_assumptions(wb)
    headcount_refs = build_headcount(wb, assumptions_refs)
    build_subsidiary_pl(wb, "USA PL", "Lumin Light USA", assumptions_refs, headcount_refs)
    build_subsidiary_pl(wb, "Nigeria PL", "Lumin Light Nigeria", assumptions_refs, headcount_refs)
    build_consolidated_pl(wb)
    build_cash_flow(wb, assumptions_refs)

    wb.save(OUT_PATH)
    print(f"Saved: {os.path.abspath(OUT_PATH)}")


if __name__ == "__main__":
    main()
